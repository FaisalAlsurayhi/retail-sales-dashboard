from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.utils import get_column_letter

wb = Workbook()

# ── Sheet 1: Raw Data ──────────────────────────────────────────────────────────
data_sheet = wb.active
data_sheet.title = "Sales Data"

headers = ["Month", "Revenue (SAR)", "Units Sold", "Avg Order Value (SAR)", "Returns (SAR)", "Net Revenue (SAR)"]
months = ["Jan 2024","Feb 2024","Mar 2024","Apr 2024","May 2024","Jun 2024",
          "Jul 2024","Aug 2024","Sep 2024","Oct 2024","Nov 2024","Dec 2024"]
revenue =       [142000,128000,155000,167000,149000,138000,172000,185000,161000,178000,210000,245000]
units =         [284,   256,   310,   334,   298,   276,   344,   370,   322,   356,   420,   490]
returns =       [4260,  3840,  4650,  5010,  4470,  4140,  5160,  5550,  4830,  5340,  6300,  7350]

# header row
header_fill = PatternFill("solid", start_color="1F4E79")
header_font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
thin = Side(style="thin", color="CCCCCC")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for col, h in enumerate(headers, 1):
    cell = data_sheet.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", wrap_text=True)
    cell.border = border

# data rows
for row_idx, (month, rev, unit, ret) in enumerate(zip(months, revenue, units, returns), 2):
    data_sheet.cell(row=row_idx, column=1, value=month)
    data_sheet.cell(row=row_idx, column=2, value=rev)
    data_sheet.cell(row=row_idx, column=3, value=unit)
    # avg order value formula
    data_sheet.cell(row=row_idx, column=4, value=f"=B{row_idx}/C{row_idx}")
    data_sheet.cell(row=row_idx, column=5, value=ret)
    # net revenue formula
    data_sheet.cell(row=row_idx, column=6, value=f"=B{row_idx}-E{row_idx}")

# totals row
total_row = len(months) + 2
data_sheet.cell(row=total_row, column=1, value="TOTAL")
data_sheet.cell(row=total_row, column=2, value=f"=SUM(B2:B{total_row-1})")
data_sheet.cell(row=total_row, column=3, value=f"=SUM(C2:C{total_row-1})")
data_sheet.cell(row=total_row, column=4, value=f"=AVERAGE(D2:D{total_row-1})")
data_sheet.cell(row=total_row, column=5, value=f"=SUM(E2:E{total_row-1})")
data_sheet.cell(row=total_row, column=6, value=f"=SUM(F2:F{total_row-1})")

total_fill = PatternFill("solid", start_color="D6E4F0")
total_font = Font(bold=True, name="Arial", size=10)
for col in range(1, 7):
    cell = data_sheet.cell(row=total_row, column=col)
    cell.fill = total_fill
    cell.font = total_font
    cell.border = border

# alt row shading + formatting
alt_fill = PatternFill("solid", start_color="EBF3FB")
sar_fmt = '#,##0'
for row_idx in range(2, total_row):
    for col in range(1, 7):
        cell = data_sheet.cell(row=row_idx, column=col)
        cell.font = Font(name="Arial", size=10)
        cell.border = border
        cell.alignment = Alignment(horizontal="center")
        if row_idx % 2 == 0:
            cell.fill = alt_fill
        if col in (2, 4, 5, 6):
            cell.number_format = sar_fmt
        if col == 3:
            cell.number_format = '#,##0'

# column widths
col_widths = [12, 18, 13, 22, 15, 18]
for col, w in enumerate(col_widths, 1):
    data_sheet.column_dimensions[get_column_letter(col)].width = w

data_sheet.row_dimensions[1].height = 32
data_sheet.freeze_panes = "A2"

# ── Sheet 2: Dashboard ─────────────────────────────────────────────────────────
dash = wb.create_sheet("Dashboard")
dash.sheet_view.showGridLines = False

# Title
dash["B2"] = "Eastern Province Retail — 2024 Sales Dashboard"
dash["B2"].font = Font(bold=True, size=16, color="1F4E79", name="Arial")
dash["B3"] = "Al-Khobar Branch  |  Annual Performance Summary"
dash["B3"].font = Font(size=11, color="666666", name="Arial", italic=True)

# KPI section title
dash["B5"] = "KEY PERFORMANCE INDICATORS"
dash["B5"].font = Font(bold=True, size=10, color="FFFFFF", name="Arial")
dash["B5"].fill = PatternFill("solid", start_color="1F4E79")
dash.merge_cells("B5:I5")
dash["B5"].alignment = Alignment(horizontal="center")

# KPI cards: label, formula pulling from Sales Data sheet
kpis = [
    ("Total Revenue",    "='Sales Data'!B14",   "SAR",  "2F75BE"),
    ("Net Revenue",      "='Sales Data'!F14",   "SAR",  "2E7D32"),
    ("Units Sold",       "='Sales Data'!C14",   "units","F57C00"),
    ("Best Month",       "=INDEX('Sales Data'!A2:A13,MATCH(MAX('Sales Data'!B2:B13),'Sales Data'!B2:B13,0))", "", "6A1B9A"),
]

kpi_cols = [2, 4, 6, 8]  # B, D, F, H
for (label, formula, unit, color), col in zip(kpis, kpi_cols):
    # label
    lbl = dash.cell(row=7, column=col, value=label)
    lbl.font = Font(bold=True, size=10, color="FFFFFF", name="Arial")
    lbl.fill = PatternFill("solid", start_color=color)
    lbl.alignment = Alignment(horizontal="center", vertical="center")
    dash.merge_cells(start_row=7, start_column=col, end_row=7, end_column=col+1)

    # value
    val = dash.cell(row=8, column=col, value=formula)
    val.font = Font(bold=True, size=18, color=color, name="Arial")
    val.alignment = Alignment(horizontal="center", vertical="center")
    dash.merge_cells(start_row=8, start_column=col, end_row=8, end_column=col+1)
    if unit == "SAR":
        val.number_format = '#,##0 "SAR"'
    elif unit == "units":
        val.number_format = '#,##0 "units"'

    # bottom border accent
    accent = dash.cell(row=9, column=col)
    accent.fill = PatternFill("solid", start_color=color)
    dash.merge_cells(start_row=9, start_column=col, end_row=9, end_column=col+1)

    # set row/col dimensions
    dash.row_dimensions[7].height = 22
    dash.row_dimensions[8].height = 38
    dash.row_dimensions[9].height = 5
    dash.column_dimensions[get_column_letter(col)].width = 16
    dash.column_dimensions[get_column_letter(col+1)].width = 3

# ── Bar Chart: Monthly Revenue ─────────────────────────────────────────────────
bar = BarChart()
bar.type = "col"
bar.title = "Monthly Revenue (SAR)"
bar.y_axis.title = "SAR"
bar.x_axis.title = "Month"
bar.style = 10
bar.width = 18
bar.height = 11

data_ref = Reference(data_sheet, min_col=2, max_col=2, min_row=1, max_row=13)
cats = Reference(data_sheet, min_col=1, min_row=2, max_row=13)
bar.add_data(data_ref, titles_from_data=True)
bar.set_categories(cats)
bar.series[0].graphicalProperties.solidFill = "2F75BE"

dash.add_chart(bar, "B11")

# ── Line Chart: Net Revenue trend ─────────────────────────────────────────────
line = LineChart()
line.title = "Net Revenue Trend (SAR)"
line.y_axis.title = "SAR"
line.x_axis.title = "Month"
line.style = 10
line.width = 18
line.height = 11

net_ref = Reference(data_sheet, min_col=6, max_col=6, min_row=1, max_row=13)
line.add_data(net_ref, titles_from_data=True)
line.set_categories(cats)
line.series[0].graphicalProperties.line.solidFill = "2E7D32"
line.series[0].graphicalProperties.line.width = 25000

dash.add_chart(line, "J11")

# ── Footer ─────────────────────────────────────────────────────────────────────
dash["B30"] = "Data source: Internal POS system  |  Built by Faisal Alsurayhi  |  github.com/faisalalsurayhi"
dash["B30"].font = Font(size=9, color="999999", name="Arial", italic=True)

wb.save("/home/claude/retail_sales_dashboard.xlsx")
print("Saved.")
